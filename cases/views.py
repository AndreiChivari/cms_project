import os
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from urllib3 import request
from .models import Dosar, ParteImplicata, Infractiune, MasuraPreventiva, IstoricDesemnare, StadiuCercetare, SolutieDosar, Notificare, TermenProcedural
from documents.forms import DocumentForm 
from .forms import DosarForm, ParteImplicataForm, CreareDosarForm, InfractiuneForm, MasuraPreventivaForm, StadiuCercetareForm, SolutieDosarForm, TermenProceduralForm
from documents.models import ActUrmarire
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.core.paginator import Paginator
from .utils import render_to_pdf
from django.http import HttpResponse
from django.conf import settings 
from datetime import datetime, date, timedelta # calculul alertelor
from django.urls import reverse # link-urile automate către dosare
from django.http import JsonResponse # folosim JavaScript pentru a sterge notificarile fara a da click pe ele si a nu reîncărca pagina
from django.contrib.auth import get_user_model
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from django.contrib import messages
import json
# Importuri pentru ocr
import pytesseract
from PIL import Image
from django.views.decorators.csrf import csrf_exempt
from PIL import Image, ImageEnhance, ImageFilter
import re
# Importuri pentru criptare CNP
import hmac
import hashlib
# Import pentru generare şi salvare documente Word
from docxtpl import DocxTemplate
from django.core.files.base import ContentFile
# --- IMPORTURI NOI PENTRU SEMNĂTURĂ ---
from pyhanko.sign import signers
from pyhanko.pdf_utils.reader import PdfFileReader
from pyhanko.pdf_utils.incremental_writer import IncrementalPdfFileWriter
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import serialization
from cryptography import x509
from asn1crypto import pem
from asn1crypto import x509 as asn1_x509
from asn1crypto import pem
from asn1crypto import x509 as asn1_x509
from asn1crypto import keys as asn1_keys  # <-- Importul nou
from pyhanko.sign.fields import SigFieldSpec, append_signature_field
from pyhanko.stamp import text
import logging
from django.utils import timezone

logger = logging.getLogger(__name__)

User = get_user_model()

from django.shortcuts import render
from django.db.models import Q, Count
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from .models import Dosar, MasuraPreventiva
from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def dashboard(request):
    utilizator = request.user
    
    # 1. STATISTICI GLOBALE
    total_dosare = Dosar.objects.count()
    
    # 2. STATISTICI PERSONALE
    if utilizator.is_superuser or getattr(utilizator, 'rol', '') == 'ADMIN':
        dosare_mele = total_dosare
        dosare_solutionate = Dosar.objects.filter(stadii_cercetare__solutii__este_finala=True).distinct().count()
        dosare_in_lucru = total_dosare - dosare_solutionate
        
        dosarele_mele_lista = Dosar.objects.prefetch_related(
            'infractiuni', 
            'stadii_cercetare__solutii'
        ).all().order_by('-data_inregistrarii')[:5]
    else:
        # Condiția pentru a găsi dosarele unde userul este implicat
        conditie_mea = Q(ofiter_caz=utilizator) | Q(procuror_caz=utilizator) | Q(grefier_caz=utilizator)
        
        # Total dosare personale
        dosare_mele = Dosar.objects.filter(conditie_mea).count()
        
        # Dosare personale SOLUȚIONATE (au măcar o soluție finală)
        dosare_solutionate = Dosar.objects.filter(
            conditie_mea, 
            stadii_cercetare__solutii__este_finala=True
        ).distinct().count()
        
        # Dosare personale ÎN LUCRU
        dosare_in_lucru = dosare_mele - dosare_solutionate
        
        # Ultimele 5 dosare pentru tabel
        dosarele_mele_lista = Dosar.objects.prefetch_related(
            'infractiuni', 
            'stadii_cercetare__solutii'
        ).filter(conditie_mea).order_by('-data_inregistrarii')[:5]

    # 3. INDICATOR DE OPERATIVITATE / ÎNCĂRCĂTURĂ
    # Calculăm câți utilizatori activi au dosare pentru a afla media
    utilizatori_activi = User.objects.filter(
        Q(dosare_instrumentate__isnull=False) | 
        Q(dosare_supravegheate__isnull=False) | 
        Q(dosare_gestionate__isnull=False)
    ).distinct().count()
    
    media_sistem = total_dosare / utilizatori_activi if utilizatori_activi > 0 else 0
    diferenta_medie = dosare_mele - media_sistem
    
    # 4. AGREGARE ALERTE (TIMELINE INTELIGENT)
    ZILE_IN_AVANS = 30  # Cu câte zile înainte apare alerta
    ZILE_TOLERANTA = 3  # Câte zile după expirare mai stă pe ecran
    azi = date.today()
    prag_viitor = azi + timedelta(days=ZILE_IN_AVANS)
    prag_trecut = azi - timedelta(days=ZILE_TOLERANTA)
    
    # Pregătim condiția pentru utilizatorii normali (să vadă doar alertele din dosarele lor)
    conditie_alerte_user = Q(dosar__ofiter_caz=utilizator) | Q(dosar__procuror_caz=utilizator) | Q(dosar__grefier_caz=utilizator)

    timeline_alerte = []

    # --- 4.1. Extragem Măsurile Preventive ---
    masuri_qs = MasuraPreventiva.objects.filter(
        data_sfarsit__lte=prag_viitor,
        data_sfarsit__gte=prag_trecut # Arată și ce a expirat în ultimele zile setate
    )
    # Filtrăm pe roluri
    if not (utilizator.is_superuser or getattr(utilizator, 'rol', '') == 'ADMIN'):
        masuri_qs = masuri_qs.filter(conditie_alerte_user)

    for m in masuri_qs:
        timeline_alerte.append({
            'tip_alerta': 'MASURA',
            'titlu': f"Măsură: {m.get_tip_masura_display()}",
            'subtitlu': f"Parte: {m.parte.nume_complet}",
            'dosar': m.dosar,
            'data_limita': m.data_sfarsit,
            'zile_ramase': m.zile_ramase,
            'expirat': m.zile_ramase < 0
        })

    # --- 4.2. Extragem Termenele Procedurale ---
    termene_qs = TermenProcedural.objects.filter(
        data_limita__lte=prag_viitor,
        data_limita__gte=prag_trecut
    )
    # Filtrăm pe roluri
    if not (utilizator.is_superuser or getattr(utilizator, 'rol', '') == 'ADMIN'):
        termene_qs = termene_qs.filter(conditie_alerte_user)

    for t in termene_qs:
        timeline_alerte.append({
            'tip_alerta': 'TERMEN',
            'titlu': f"Termen: {t.get_tip_termen_display()}",
            'subtitlu': t.detalii if t.detalii else "Atenție la termenul procedural!",
            'dosar': t.dosar,
            'data_limita': t.data_limita,
            'zile_ramase': t.zile_ramase,
            'expirat': t.zile_ramase < 0
        })

    # --- 4.3. Unificare și Ordonare ---
    # Ordonăm toată lista combinată în funcție de zilele rămase (cele mai puține zile primele)
    timeline_alerte = sorted(timeline_alerte, key=lambda x: x['zile_ramase'])
    
    # Păstrăm doar primele 10 urgențe ca să nu încărcăm vizual prea mult dashboard-ul
    timeline_alerte = timeline_alerte[:10]

# ==========================================
    # 4.5. DATE PENTRU GRAFICUL CU BARE (Ultimele 6 luni)
    # ==========================================
    luni_ro = {1: 'Ian', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Mai', 6: 'Iun', 
               7: 'Iul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    
    labels_grafic = []
    date_grafic = []
    
    # Ne setăm ca punct de pornire data de 1 a lunii curente
    luna_start = azi.replace(day=1)

    # Mergem 5 luni în spate + luna curentă = 6 luni
    for i in range(11, -1, -1):
        an_calcul = luna_start.year
        luna_calcul = luna_start.month - i
        
        # Dacă luna scade sub 1 (ex. suntem în Februarie și mergem 3 luni în spate), ajustăm anul
        if luna_calcul <= 0:
            an_calcul -= 1
            luna_calcul += 12
            
        nume_eticheta = f"{luni_ro[luna_calcul]} {an_calcul}"
        labels_grafic.append(nume_eticheta)
        
        # Numărăm dosarele pentru acea lună și an
        if utilizator.is_superuser or getattr(utilizator, 'rol', '') == 'ADMIN':
            nr_dosare = Dosar.objects.filter(data_inregistrarii__year=an_calcul, data_inregistrarii__month=luna_calcul).count()
        else:
            nr_dosare = Dosar.objects.filter(conditie_mea, data_inregistrarii__year=an_calcul, data_inregistrarii__month=luna_calcul).count()
            
        date_grafic.append(nr_dosare)

    # 5. CONTEXT PENTRU HTML
    context = {
        'user_rol': getattr(utilizator, 'rol', 'Membru Echipă'),
        # Stats
        'total_dosare': total_dosare,
        'dosare_mele': dosare_mele,
        'dosare_solutionate': dosare_solutionate,
        'dosare_in_lucru': dosare_in_lucru,
        # Operativitate
        'media_sistem': round(media_sistem, 1),
        'diferenta_medie': round(diferenta_medie, 1),
        # Procente pentru grafic
        'procent_lucru': int((dosare_in_lucru / dosare_mele * 100)) if dosare_mele > 0 else 0,
        'procent_solutionate': int((dosare_solutionate / dosare_mele * 100)) if dosare_mele > 0 else 0,
        # Liste
        'dosarele_mele_lista': dosarele_mele_lista,
        
        # Variabila NOASTRĂ NOUĂ unificată
        'timeline_alerte': timeline_alerte,

        # DATE NOI PENTRU GRAFIC (transformate în format JSON ca să le înțeleagă JavaScript-ul)
        'labels_grafic': json.dumps(labels_grafic),
        'date_grafic': json.dumps(date_grafic),
    }
    
    return render(request, 'cases/dashboard.html', context)

@login_required
def adaugare_dosar(request):
    if request.method == 'POST':
        form = CreareDosarForm(request.POST)
        if form.is_valid():
            dosar = form.save()

            # Creăm istoricul inițial
            if dosar.ofiter_caz:
                IstoricDesemnare.objects.create(dosar=dosar, utilizator=dosar.ofiter_caz, rol='Ofițer', data_desemnare=dosar.data_inregistrarii)
            if dosar.procuror_caz:
                IstoricDesemnare.objects.create(dosar=dosar, utilizator=dosar.procuror_caz, rol='Procuror', data_desemnare=dosar.data_inregistrarii)
            if dosar.grefier_caz:
                IstoricDesemnare.objects.create(dosar=dosar, utilizator=dosar.grefier_caz, rol='Grefier', data_desemnare=dosar.data_inregistrarii)

            # Notificări la Crearea Dosarului
            link_dosar = reverse('cases:detalii_dosar', args=[dosar.pk])
            mesaj_nou = f"Ai fost desemnat pe dosarul nou {dosar.numar_unic}."
            
            if dosar.ofiter_caz:
                Notificare.objects.create(utilizator=dosar.ofiter_caz, mesaj=mesaj_nou, link=link_dosar)
            if dosar.procuror_caz:
                Notificare.objects.create(utilizator=dosar.procuror_caz, mesaj=mesaj_nou, link=link_dosar)
            if dosar.grefier_caz:
                Notificare.objects.create(utilizator=dosar.grefier_caz, mesaj=mesaj_nou, link=link_dosar)

            return redirect('cases:detalii_dosar', pk=dosar.pk)
    else:
        form = CreareDosarForm()
        
    context = {'form': form}
    return render(request, 'cases/adaugare_dosar.html', context)

@login_required
@login_required
def lista_dosare(request):
    # 1. Luăm inputul utilizatorului
    query_text = request.GET.get('q', '').strip() 
    stadiu_filtru = request.GET.get('stadiu', '')

    context = {
        'query_text': query_text,
        'stadiu_filtru': stadiu_filtru,
    }

    # MODULUL DE CĂUTARE GLOBALĂ (Dacă utilizatorul a căutat ceva)
    if query_text:
        # TRUC PENTRU DIACRITICE ÎN SQLITE:
        # Generăm 3 variante ale textului pentru a ocoli limitarea SQLite privind majusculele Unicode (ă, î, ș, ț, â)
        q_lower = query_text.lower()
        q_upper = query_text.upper()
        q_title = query_text.title()

        # 1. Căutăm cuvântul în etichetele Actelor Normative și extragem CHEILE (ex: 'CP')
        chei_acte_normative = [
            cheie for cheie, eticheta in Infractiune.ActNormativ.choices
            if q_lower in eticheta.lower()
        ]
        
        # 2. Căutăm cuvântul în etichetele Măsurilor Preventive și extragem CHEILE (ex: 'AREST_PREVENTIV')
        chei_masuri = [
            cheie for cheie, eticheta in MasuraPreventiva.TipMasura.choices
            if q_lower in eticheta.lower()
        ]

        # A. Căutăm direct în tabelul Dosare
        dosare_gasite = Dosar.objects.filter(
            Q(numar_unic__icontains=query_text) | 
            Q(infractiune_cercetata__icontains=q_lower) | 
            Q(infractiune_cercetata__icontains=q_upper) | 
            Q(infractiune_cercetata__icontains=q_title)
        ).distinct()

        # B. Căutăm prin modelul de Infracțiune
        dosare_din_infractiuni = Dosar.objects.filter(
            Q(infractiuni__adresa_comiterii__icontains=q_lower) |
            Q(infractiuni__adresa_comiterii__icontains=q_upper) |
            Q(infractiuni__adresa_comiterii__icontains=q_title) |
            
            Q(infractiuni__incadrare_juridica__icontains=q_lower) |
            Q(infractiuni__incadrare_juridica__icontains=q_upper) |
            Q(infractiuni__incadrare_juridica__icontains=q_title) |
            
            Q(infractiuni__articol__icontains=query_text) |
            Q(infractiuni__act_normativ__in=chei_acte_normative) # <--- Găsește dosarele după cod (ex: "penal" -> CP)
        ).distinct()
        
        toate_dosarele_gasite = (dosare_gasite | dosare_din_infractiuni).distinct().order_by('-data_inregistrarii')
        
        # C. Căutăm în Părți Implicate
        # Filtrul de bază (căutarea în textele normale)
        filtru_persoane = (
            Q(nume_complet__icontains=q_lower) |
            Q(nume_complet__icontains=q_upper) |
            Q(nume_complet__icontains=q_title) |
            Q(adresa__icontains=q_lower) |
            Q(adresa__icontains=q_upper) |
            Q(adresa__icontains=q_title) |
            Q(mentiuni__icontains=q_lower) |
            Q(mentiuni__icontains=q_upper) |
            Q(mentiuni__icontains=q_title)
        )

        # Verificăm dacă utilizatorul a introdus exact un CNP (13 cifre)
        if len(query_text) == 13 and query_text.isdigit():
            # Calculăm hash-ul exact cum o face modelul la salvare
            secret = settings.SECRET_KEY.encode('utf-8')
            hash_cautat = hmac.new(secret, query_text.encode('utf-8'), hashlib.sha256).hexdigest()
            
            # Adăugăm (prin operatorul OR " | ") condiția de căutare exactă a hash-ului
            filtru_persoane = filtru_persoane | Q(cnp_hash=hash_cautat)

        # Aplicăm filtrul pe baza de date
        persoane_gasite = ParteImplicata.objects.filter(
            filtru_persoane
        ).select_related('dosar').order_by('nume_complet')[:20]  # Adăugat [:20] pentru a limita afișarea și a preveni blocajele

        # D. Căutăm în Documente/Acte Urmărire
        documente_gasite = ActUrmarire.objects.filter(
            Q(titlu__icontains=q_lower) |
            Q(titlu__icontains=q_upper) |
            Q(titlu__icontains=q_title) |
            Q(descriere_scurta__icontains=q_lower) |
            Q(descriere_scurta__icontains=q_upper) |
            Q(descriere_scurta__icontains=q_title)
        ).select_related('dosar').order_by('-data_incarcarii')

        # E. Căutăm în Măsuri Preventive
        masuri_gasite = MasuraPreventiva.objects.filter(
            Q(tip_masura__in=chei_masuri) | 
            Q(parte__nume_complet__icontains=q_lower) | 
            Q(parte__nume_complet__icontains=q_upper) |
            Q(parte__nume_complet__icontains=q_title)
        ).select_related('dosar', 'parte').order_by('-data_inceput')

        # Trimitem rezultatele către HTML
        context.update({
            'mod_cautare': True,
            'dosare_gasite': toate_dosarele_gasite,
            'persoane_gasite': persoane_gasite,
            'documente_gasite': documente_gasite,
            'masuri_gasite': masuri_gasite, # <--- Am adăugat noua listă
            'total_rezultate': toate_dosarele_gasite.count() + persoane_gasite.count() + documente_gasite.count() + masuri_gasite.count()
        })

    # MODUL IMPLICIT (Afișarea Listei, dacă utilizatorul nu a căutat nimic)
    else:
        dosare = Dosar.objects.all().order_by('-data_inregistrarii')
        
        if stadiu_filtru:
            pass 

        paginator = Paginator(dosare, 10) # dosare afişate pe pagină
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)  

        context.update({
            'mod_cautare': False,
            'page_obj': page_obj,
        })
        
    return render(request, 'cases/lista_dosare.html', context)

@login_required
def detalii_dosar(request, pk):
    dosar = get_object_or_404(Dosar, pk=pk)
    
    # Calculăm o singură dată dacă are drepturi, pentru a trimite către HTML
    poate_edita = dosar.are_drepturi_editare(request.user)
    
    form_document = DocumentForm()
    form_parte = ParteImplicataForm()
    form_infractiune = InfractiuneForm()
    form_termen = TermenProceduralForm() 

    # 1. Inițializăm formularul NOU, dându-i ID-ul dosarului
    form_masura = MasuraPreventivaForm(dosar_id=dosar.pk)
    
    if request.method == 'POST':
        if not poate_edita:
            raise PermissionDenied("Nu ai permisiunea de a modifica acest dosar.")
        # VERIFICARE 1: A apăsat butonul de adăugare Document?
        if 'btn_salveaza_document' in request.POST:
            form_document = DocumentForm(request.POST, request.FILES)
            if form_document.is_valid():
                document = form_document.save(commit=False)
                document.dosar = dosar
                document.autor = request.user
                document.save()
                return redirect('cases:detalii_dosar', pk=dosar.pk)
                
        # VERIFICARE 2: A apăsat butonul de adăugare Parte Implicată?
        elif 'btn_salveaza_parte' in request.POST:
            form_parte = ParteImplicataForm(request.POST)
            if form_parte.is_valid():
                # 1. Salvăm persoana în dosar
                parte = form_parte.save(commit=False)
                parte.dosar = dosar
                parte.save()
                
                # 2. Integrare între module: Salvăm poza ca Document
                salveaza_doc = request.POST.get('salveaza_ca_document')
                fisier_ci = request.FILES.get('fisier_copie_ci')
                
                if salveaza_doc == 'DA' and fisier_ci:
                    ActUrmarire.objects.create(
                        titlu=f"Copie C.I. - {parte.nume_complet}",
                        tip=ActUrmarire.TipDocument.ALTUL, # Folosim tipul 'ALTUL'
                        dosar=dosar,
                        autor=request.user,
                        fisier=fisier_ci,
                        descriere_scurta="Document generat automat la scanarea C.I. (Modul OCR)."
                    )

                return redirect('cases:detalii_dosar', pk=dosar.pk)
            
        # 2. LOGICA PENTRU SALVAREA INFRACȚIUNII
        elif 'btn_salveaza_infractiune' in request.POST:
            form_infractiune = InfractiuneForm(request.POST)
            if form_infractiune.is_valid():
                infractiune = form_infractiune.save(commit=False)
                infractiune.dosar = dosar  # Legăm infracțiunea de dosarul curent!
                infractiune.save()
                return redirect('cases:detalii_dosar', pk=dosar.pk)
            
        # 2. Logica pentru Măsuri Preventive
        elif 'btn_salveaza_masura' in request.POST:
            # Citim datele trimise și îi dăm din nou ID-ul dosarului pentru validare
            form_masura = MasuraPreventivaForm(request.POST, dosar_id=dosar.pk)
            if form_masura.is_valid():
                masura = form_masura.save(commit=False)
                masura.dosar = dosar
                masura.save()
                return redirect('cases:detalii_dosar', pk=dosar.pk)
            
        # 2. Dacă a apăsat pe Salvare TERMEN PROCEDURAL (Cod Nou)
        elif 'btn_salveaza_termen' in request.POST:
            form_termen = TermenProceduralForm(request.POST)
            if form_termen.is_valid():
                termen = form_termen.save(commit=False)
                termen.dosar = dosar
                termen.save()
                messages.success(request, 'Termenul procedural a fost adăugat cu succes.')
                return redirect('cases:detalii_dosar', pk=dosar.pk)

    context = {
            'dosar': dosar,
            'form_document': form_document,
            'form_parte': form_parte,
            'form_infractiune': form_infractiune,
            'form_masura': form_masura,
            'form_termen': form_termen,
            'poate_edita': poate_edita 
        }
    
    return render(request, 'cases/detalii_dosar.html', context)

@login_required
def editare_dosar(request, pk):
    # Găsim dosarul pe care vrem să-l edităm
    dosar = get_object_or_404(Dosar, pk=pk)

    # 1. Memorăm echipa VECHE înainte să fie modificată
    vechi_ofiter = dosar.ofiter_caz
    vechi_procuror = dosar.procuror_caz
    vechi_grefier = dosar.grefier_caz

    # SECURITATE:
    if not dosar.are_drepturi_editare(request.user):
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")
    
    if request.method == 'POST':
        # request.POST conține noile date. 
        # instance=dosar încarcă dosarul existent în formular, iar request.POST îl actualizează cu noile valori.
        form = DosarForm(request.POST, instance=dosar)
        if form.is_valid():
            # Preluăm data din noul câmp. Dacă grutilizatorul nu a trecut-o punem automat data de azi.
            data_schimbare = form.cleaned_data.get('data_schimbare_echipa') or date.today()
            
            nou_dosar = form.save() # Salvăm dosarul cu noua echipă

            # Funcție internă care face schimbul pentru fiecare rol
            def actualizeaza_istoric(rol_nume, vechi_user, nou_user):
                if vechi_user != nou_user:
                    # Dacă exista cineva înainte, îi închidem mandatul
                    if vechi_user:
                        IstoricDesemnare.objects.filter(
                            dosar=nou_dosar, utilizator=vechi_user, rol=rol_nume, data_finalizare__isnull=True
                        ).update(data_finalizare=data_schimbare)
                    
                    # Creăm mandatul nou pentru noul venit
                    if nou_user:
                        IstoricDesemnare.objects.create(
                            dosar=nou_dosar, utilizator=nou_user, rol=rol_nume, data_desemnare=data_schimbare
                        )

                        # Notificare pentru noul membru 
                        link_dosar = reverse('cases:detalii_dosar', args=[nou_dosar.pk])
                        Notificare.objects.create(
                            utilizator=nou_user, 
                            mesaj=f"Ai preluat mandatul de {rol_nume} pe dosarul {nou_dosar.numar_unic}.", 
                            link=link_dosar
                        )

            # 3. Executăm verificarea pentru toți 3:
            actualizeaza_istoric('Ofițer', vechi_ofiter, nou_dosar.ofiter_caz)
            actualizeaza_istoric('Procuror', vechi_procuror, nou_dosar.procuror_caz)
            actualizeaza_istoric('Grefier', vechi_grefier, nou_dosar.grefier_caz)
            
            # După salvare, trimitem utilizatorul înapoi pe pagina de detalii a dosarului
            return redirect('cases:detalii_dosar', pk=dosar.pk)
    else:
        # Dacă doar accesăm pagina, populăm formularul cu datele existente ale dosarului
        form = DosarForm(instance=dosar)
        
    context = {
        'form': form,
        'dosar': dosar
    }
    return render(request, 'cases/editare_dosar.html', context)

@login_required
def stergere_istoric_echipa(request, pk):
    # 1. Preluăm înregistrarea din istoric
    istoric = get_object_or_404(IstoricDesemnare, pk=pk)
    
    # 2. DEFINIM VARIABILA DOSAR (obiectul întreg, nu doar ID-ul)
    dosar = istoric.dosar
    
    # 3. SINCRONIZARE BAZĂ DE DATE
    if istoric.rol == 'Ofițer' and dosar.ofiter_caz == istoric.utilizator:
        dosar.ofiter_caz = None
        dosar.save(update_fields=['ofiter_caz'])
        
    elif istoric.rol == 'Procuror' and dosar.procuror_caz == istoric.utilizator:
        dosar.procuror_caz = None
        dosar.save(update_fields=['procuror_caz'])
        
    elif istoric.rol == 'Grefier' and dosar.grefier_caz == istoric.utilizator:
        dosar.grefier_caz = None
        dosar.save(update_fields=['grefier_caz'])

    # 4. Salvăm ID-ul dosarului separat, pentru că pe rândul următor ștergem istoricul
    dosar_id = dosar.pk 
    
    # 5. Ștergem înregistrarea din istoric
    istoric.delete()
    messages.success(request, "Înregistrarea a fost ștearsă din istoric cu succes!")
    
    # 6. Ne întoarcem la dosar
    return redirect('cases:detalii_dosar', pk=dosar_id)

@login_required
def editare_parte(request, pk):
    # Găsim persoana după ID
    parte = get_object_or_404(ParteImplicata, pk=pk)
    # Reținem ID-ul dosarului pentru a ști unde să ne întoarcem
    dosar_id = parte.dosar.pk 

    # SECURITATE:
    if not parte.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")
    
    if request.method == 'POST':
        # Punem instance=parte pentru a suprascrie datele existente, nu a crea una nouă
        form = ParteImplicataForm(request.POST, instance=parte)
        if form.is_valid():
            form.save()
            return redirect('cases:detalii_dosar', pk=dosar_id)
    else:
        form = ParteImplicataForm(instance=parte)
        
    context = {'form': form, 'parte': parte}
    return render(request, 'cases/editare_parte.html', context)

@login_required
def stergere_parte(request, pk):
    parte = get_object_or_404(ParteImplicata, pk=pk)
    dosar_id = parte.dosar.pk
    
    # Verificăm drepturile pe dosar
    if not parte.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")

    # Pentru ștergere, este o bună practică să cerem confirmare printr-un formular POST
    if request.method == 'POST':
        parte.delete() # Șterge din baza de date
        return redirect('cases:detalii_dosar', pk=dosar_id)
        
    context = {'parte': parte}
    return render(request, 'cases/stergere_parte.html', context)

@login_required
def editare_document(request, pk):
    document = get_object_or_404(ActUrmarire, pk=pk)
    dosar_id = document.dosar.pk

    if not document.are_drepturi_editare(request.user):
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")
    
    if request.method == 'POST':
        # request.FILES este obligatoriu aici pentru a putea schimba PDF-ul/Word-ul
        form = DocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            form.save()
            return redirect('cases:detalii_dosar', pk=dosar_id)
    else:
        form = DocumentForm(instance=document)
        
    context = {'form': form, 'document': document}
    return render(request, 'cases/editare_document.html', context)

@login_required
def stergere_document(request, pk):
    document = get_object_or_404(ActUrmarire, pk=pk)
    dosar_id = document.dosar.pk

    if not document.are_drepturi_editare(request.user):
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")
     
    if request.method == 'POST':
        # șteargem și fișierul fizic, dacă există
        if document.fisier:
            document.fisier.delete(save=False)
            
        document.delete() # Șterge înregistrarea din tabel
        return redirect('cases:detalii_dosar', pk=dosar_id)
        
    context = {'document': document}
    return render(request, 'cases/stergere_document.html', context)

@login_required
def generare_pdf_dosar(request, pk):
    dosar = get_object_or_404(Dosar, pk=pk)
    
    context = {
        'dosar': dosar,
        'stadiu_curent': dosar.stadiu_curent,
        'solutie_curenta': dosar.solutie_curenta,
        'request': request,
    }
    
    # 3. Generăm PDF-ul
    pdf = render_to_pdf('cases/pdf_template.html', context)
    
    if pdf:
        nume_fisier = f"Fisa_Dosar_{dosar.numar_unic.replace('/', '_')}.pdf"
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nume_fisier}"'
        return response
        
    return HttpResponse("A apărut o eroare la generarea PDF-ului.", status=400)

@login_required
def stergere_masura(request, pk):
    # Găsim măsura în baza de date
    masura = get_object_or_404(MasuraPreventiva, pk=pk)
    dosar_id = masura.dosar.pk
    
    # Verificăm dacă utilizatorul are drepturi pe dosarul respectiv
    if not masura.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a șterge date din acest dosar.")
        
    # Dacă utilizatorul a apăsat "Da, șterge"
    if request.method == 'POST':
        masura.delete()
        return redirect('cases:detalii_dosar', pk=dosar_id)
        
    # Dacă a dat doar click pe link, îi afișăm pagina de confirmare
    context = {
        'masura': masura,
        'dosar': masura.dosar
    }
    return render(request, 'cases/stergere_masura.html', context)

@login_required
def editare_masura(request, pk):
    masura = get_object_or_404(MasuraPreventiva, pk=pk)
    
    # Securitate: Verificăm drepturile dosarului părinte
    if not masura.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a edita date din acest dosar.")
        
    if request.method == 'POST':
        # Trimitem și `dosar_id` pentru ca formularul să știe să filtreze lista de persoane
        form = MasuraPreventivaForm(request.POST, instance=masura, dosar_id=masura.dosar.pk)
        if form.is_valid():
            form.save()
            return redirect('cases:detalii_dosar', pk=masura.dosar.pk)
    else:
        # Pre-completăm formularul cu datele existente
        form = MasuraPreventivaForm(instance=masura, dosar_id=masura.dosar.pk)
        
    context = {
        'form': form,
        'masura': masura,
        'dosar': masura.dosar
    }
    return render(request, 'cases/editare_masura.html', context)

@login_required
def editare_termen(request, pk):
    termen = get_object_or_404(TermenProcedural, pk=pk)
    dosar = termen.dosar
    
    if not termen.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a edita date din acest dosar.")

    if request.method == 'POST':
        form = TermenProceduralForm(request.POST, instance=termen)
        if form.is_valid():
            form.save()
            messages.success(request, 'Termenul procedural a fost actualizat cu succes.')
            return redirect('cases:detalii_dosar', pk=dosar.pk)
    else:
        form = TermenProceduralForm(instance=termen)

    context = {
        'form': form,
        'termen': termen,
        'dosar': dosar
    }
    return render(request, 'cases/editare_termen.html', context)

@login_required
def stergere_termen(request, pk):
    termen = get_object_or_404(TermenProcedural, pk=pk)
    dosar = termen.dosar

    if not termen.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a șterge date din acest dosar.")

    if request.method == 'POST':
        termen.delete()
        messages.success(request, 'Termenul procedural a fost șters definitiv.')
        return redirect('cases:detalii_dosar', pk=dosar.pk)

    context = {
        'termen': termen,
        'dosar': dosar
    }
    return render(request, 'cases/stergere_termen.html', context)

@login_required
def editare_infractiune(request, pk):
    infractiune = get_object_or_404(Infractiune, pk=pk)
    
    if not infractiune.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a edita date din acest dosar.")
        
    if request.method == 'POST':
        form = InfractiuneForm(request.POST, instance=infractiune)
        if form.is_valid():
            form.save()
            return redirect('cases:detalii_dosar', pk=infractiune.dosar.pk)
    else:
        form = InfractiuneForm(instance=infractiune)
        
    context = {
        'form': form,
        'infractiune': infractiune,
        'dosar': infractiune.dosar
    }
    return render(request, 'cases/editare_infractiune.html', context)

@login_required
def stergere_infractiune(request, pk):
    infractiune = get_object_or_404(Infractiune, pk=pk)
    dosar_id = infractiune.dosar.pk
    
    if not infractiune.dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a șterge date din acest dosar.")
        
    if request.method == 'POST':
        infractiune.delete()
        return redirect('cases:detalii_dosar', pk=dosar_id)
        
    context = {
        'infractiune': infractiune,
        'dosar': infractiune.dosar
    }
    return render(request, 'cases/stergere_infractiune.html', context)

@login_required
def gestionare_stadii(request, pk):
    dosar = get_object_or_404(Dosar, pk=pk)
    
    # Securitate
    if not dosar.are_drepturi_editare(request.user) and not request.user.is_superuser:
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")

    form_stadiu = StadiuCercetareForm()
    form_solutie = SolutieDosarForm()

    if request.method == 'POST':
    # Funcție internă ajutătoare pentru a nu repeta codul de trimitere
        def trimite_notificare_colaborare(mesaj_actiune):
            link_dosar = reverse('cases:detalii_dosar', args=[dosar.pk])
            destinatari = [dosar.ofiter_caz, dosar.procuror_caz, dosar.grefier_caz]
            
            for destinatar in destinatari:
                # Trimitem doar dacă postul e ocupat ȘI utilizatorul nu este cel care a făcut modificarea
                if destinatar and destinatar != request.user:
                    Notificare.objects.create(utilizator=destinatar, mesaj=mesaj_actiune, link=link_dosar)

        # 1. Salvează Stadiu
        if 'salveaza_stadiu' in request.POST:
            stadiu_id = request.POST.get('stadiu_id')
            if stadiu_id:
                stadiu = get_object_or_404(StadiuCercetare, pk=stadiu_id, dosar=dosar)
                form_stadiu = StadiuCercetareForm(request.POST, instance=stadiu)
            else:
                form_stadiu = StadiuCercetareForm(request.POST)
                form_stadiu.instance.dosar = dosar
                
            if form_stadiu.is_valid():
                stadiu_salvat = form_stadiu.save()
                
                # NOTIFICARE: Citim checkbox-ul
                if form_stadiu.cleaned_data.get('notifica_echipa'):
                    mesaj = f"{request.user.get_full_name() or request.user.username} a adăugat/modificat stadiul ({stadiu_salvat.get_tip_stadiu_display()}) pe dosarul {dosar.numar_unic}."
                    trimite_notificare_colaborare(mesaj)
                    
                return redirect('cases:gestionare_stadii', pk=dosar.pk)

        # 2. Salvează Soluție
        elif 'salveaza_solutie' in request.POST:
            solutie_id = request.POST.get('solutie_id')
            stadiu_parinte_id = request.POST.get('stadiu_parinte_id')
            
            if solutie_id:
                solutie = get_object_or_404(SolutieDosar, pk=solutie_id)
                form_solutie = SolutieDosarForm(request.POST, instance=solutie)
            else:
                form_solutie = SolutieDosarForm(request.POST)
                if stadiu_parinte_id:
                    form_solutie.instance.stadiu_id = stadiu_parinte_id
                    
            if form_solutie.is_valid():
                sol_salvata = form_solutie.save()
                
                #  NOTIFICARE: Citim checkbox-ul
                if form_solutie.cleaned_data.get('notifica_echipa'):
                    mesaj = f"{request.user.get_full_name() or request.user.username} a adăugat/modificat o soluție ({sol_salvata.get_tip_solutie_display()}) pe dosarul {dosar.numar_unic}."
                    trimite_notificare_colaborare(mesaj)
                    
                return redirect('cases:gestionare_stadii', pk=dosar.pk)

        # 3. Șterge Stadiu
        elif 'sterge_stadiu' in request.POST:
            stadiu_id = request.POST.get('stadiu_id')
            if stadiu_id:
                stadiu = get_object_or_404(StadiuCercetare, pk=stadiu_id, dosar=dosar)
                stadiu.delete()
                return redirect('cases:gestionare_stadii', pk=dosar.pk)
                
        # 4. Șterge Soluție
        elif 'sterge_solutie' in request.POST:
            solutie_id = request.POST.get('solutie_id')
            if solutie_id:
                solutie = get_object_or_404(SolutieDosar, pk=solutie_id)
                solutie.delete()
                return redirect('cases:gestionare_stadii', pk=dosar.pk)

    # Preluăm toate stadiile cu soluțiile lor atașate
    stadii = dosar.stadii_cercetare.all().prefetch_related('solutii')

    context = {
        'dosar': dosar,
        'form_stadiu': form_stadiu,
        'form_solutie': form_solutie,
        'stadii': stadii
    }
    return render(request, 'cases/gestionare_stadii.html', context)

@login_required
def citeste_notificare(request, pk):
    # Găsim notificarea (asigurându-ne că e a utilizatorului curent)
    notificare = get_object_or_404(Notificare, pk=pk, utilizator=request.user)
    # O marcăm ca citită
    notificare.citita = True
    notificare.save()
    # Îl redirecționăm către link-ul stocat (ex: /cases/4/)
    return redirect(notificare.link)

@login_required
def sterge_notificare_ajax(request, pk):
    if request.method == 'POST':
        notificare = get_object_or_404(Notificare, pk=pk, utilizator=request.user)
        notificare.citita = True
        notificare.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def generare_rapoarte(request):
    dosare = Dosar.objects.select_related('ofiter_caz', 'procuror_caz', 'grefier_caz').prefetch_related(
        'stadii_cercetare__solutii',
        'parti_implicate',
        'masuri_preventive',
        'infractiuni' 
    ).all().distinct()
    
    # 1. FILTRE (Criteriile de căutare)
    data_inreg_start = request.GET.get('data_inreg_start')
    data_inreg_end = request.GET.get('data_inreg_end')
    procuror_id = request.GET.get('procuror')
    ofiter_id = request.GET.get('ofiter')
    act_normativ = request.GET.get('act_normativ')
    articol = request.GET.get('articol')
    stadiu = request.GET.get('stadiu')
    solutie = request.GET.get('solutie')
    
    # Aplicăm filtrele
    if data_inreg_start:
        dosare = dosare.filter(data_inregistrarii__gte=data_inreg_start)
    if data_inreg_end:
        dosare = dosare.filter(data_inregistrarii__lte=data_inreg_end)
    if procuror_id:
        dosare = dosare.filter(procuror_caz_id=procuror_id)
    if ofiter_id:
        dosare = dosare.filter(ofiter_caz_id=ofiter_id)
    if act_normativ:
        dosare = dosare.filter(infractiuni__act_normativ=act_normativ)
    if articol:
        dosare = dosare.filter(infractiuni__articol__icontains=articol)
    if stadiu:
        dosare = dosare.filter(stadii_cercetare__tip_stadiu=stadiu)
    if solutie:
        dosare = dosare.filter(stadii_cercetare__solutii__tip_solutie=solutie)

    # 2. COLOANELE SELECTATE PENTRU AFIȘARE
    def is_checked(nume_camp, implicit=False):
        if not request.GET: 
            return implicit
        return request.GET.get(nume_camp) == 'on'

    coloane = {
        'col_numar': is_checked('col_numar', True),
        'col_data_inreg': is_checked('col_data_inreg', True),
        'col_situatie_fapt': is_checked('col_situatie_fapt', True), 
        'col_incadrare': is_checked('col_incadrare', True),  
        'col_ofiter': is_checked('col_ofiter', True),
        'col_procuror': is_checked('col_procuror', True),
        'col_stadiu_curent': is_checked('col_stadiu_curent', True),
        'col_data_stadiu': is_checked('col_data_stadiu', False),
        'col_solutie_finala': is_checked('col_solutie_finala', False),
        'col_data_solutie': is_checked('col_data_solutie', False),
        'col_parti': is_checked('col_parti', False),
        'col_masuri': is_checked('col_masuri', False),
    }

    # EXPORT EXCEL (.XLSX)
    if 'export_excel' in request.GET:
        # 1. Creăm un fișier Excel în memorie
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raport Dosare"
        
        def curata_text(text):
            if not text: return "-"
            return str(text).replace('\n', ' ').replace('\r', ' ').strip()
        
        # 2. Generăm și scriem Capul de Tabel (Headers)
        headers = []
        if coloane['col_numar']: headers.append("Numar unic")
        if coloane['col_data_inreg']: headers.append("Data înreg.")
        if coloane['col_situatie_fapt']: headers.append("Situatie fapt")
        if coloane['col_incadrare']: headers.append("Încadrare juridică")
        if coloane['col_ofiter']: headers.append("Poliţist")
        if coloane['col_procuror']: headers.append("Procuror")
        if coloane['col_stadiu_curent']: headers.append("Stadiu curent")
        if coloane['col_data_stadiu']: headers.append("Data stadiu")
        if coloane['col_solutie_finala']: headers.append("Solutie finală")
        if coloane['col_data_solutie']: headers.append("Data solutie")
        if coloane['col_parti']: headers.append("Părţi în cauză")
        if coloane['col_masuri']: headers.append("Măsuri preventive")
        
        ws.append(headers)
        
        # Stilizăm capul de tabel (bold și centrat)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
        # 3. Scriem datele (Rândurile)
        for dosar in dosare:
            row = []
            
            # 1. Folosim proprietățile pentru a lua mereu STAREA LA ZI
            stadiu = dosar.stadiu_curent
            solutie = dosar.solutie_curenta
            
            if coloane['col_numar']: 
                row.append(curata_text(dosar.numar_unic))
            if coloane['col_data_inreg']: 
                row.append(dosar.data_inregistrarii.strftime("%d.%m.%Y") if dosar.data_inregistrarii else "-")
            if coloane['col_situatie_fapt']: 
                row.append(curata_text(dosar.infractiune_cercetata))
                
            if coloane['col_incadrare']: 
                # 2. Construim lista de încadrări verificând prezența articolului (Fără 'None')
                lista_incadrari = []
                for i in dosar.infractiuni.all():
                    if i.articol and str(i.articol) != 'None':
                        lista_incadrari.append(f"art. {i.articol} - {i.get_act_normativ_display()}")
                    else:
                        lista_incadrari.append(i.get_act_normativ_display())
                
                incadrari = " | ".join(lista_incadrari)
                row.append(curata_text(incadrari))
                
            if coloane['col_ofiter']: 
                row.append(curata_text(dosar.ofiter_caz.get_full_name() if dosar.ofiter_caz else "-"))
            if coloane['col_procuror']: 
                row.append(curata_text(dosar.procuror_caz.get_full_name() if dosar.procuror_caz else "-"))
            if coloane['col_stadiu_curent']: 
                row.append(curata_text(stadiu.get_tip_stadiu_display() if stadiu else "Neînceput"))
            if coloane['col_data_stadiu']: 
                row.append(stadiu.data_incepere.strftime("%d.%m.%Y") if stadiu and stadiu.data_incepere else "-")
            if coloane['col_solutie_finala']: 
                row.append(curata_text(solutie.get_tip_solutie_display() if solutie else "-"))
            if coloane['col_data_solutie']: 
                row.append(solutie.data_solutiei.strftime("%d.%m.%Y") if solutie and solutie.data_solutiei else "-")
            if coloane['col_parti']: 
                parti = " | ".join([f"{p.nume_complet} ({p.get_calitate_procesuala_display()})" for p in dosar.parti_implicate.all()])
                row.append(curata_text(parti))
            if coloane['col_masuri']: 
                masuri = " | ".join([f"{m.get_tip_masura_display()} (exp. {m.data_sfarsit.strftime('%d.%m.%Y')})" for m in dosar.masuri_preventive.all() if m.data_sfarsit])
                row.append(curata_text(masuri))
                
            ws.append(row)
            
        # 4. Ajustăm lățimea coloanelor automat
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60) # Limităm lățimea maximă la 60
            ws.column_dimensions[column_letter].width = adjusted_width

        # 5. Salvăm în memoria serverului
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        
        # 6. Trimitem fișierul către browser cu formatul corect de .xlsx
        response = HttpResponse(
            virtual_workbook, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Raport_Dosare_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
        
        return response

    # PAGINARE: Pentru a păstra toate celelalte query params (filtrele și coloanele selectate) în link-urile de paginare, le reconstruim manual:
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    # 2. Împărțim rezultatele (câte 20 de dosare pe pagină)
    paginator = Paginator(dosare, 20) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Extragem opțiunile din modele pentru a popula automat dropdown-urile din HTML
    context = {
        'dosare': page_obj,
        'coloane': coloane,
        'lista_procurori': User.objects.filter(rol='PROCUROR'),
        'lista_ofiteri': User.objects.filter(rol='POLITIST'),
        'acte_normative': Infractiune._meta.get_field('act_normativ').choices,
        'stadii_choices': StadiuCercetare._meta.get_field('tip_stadiu').choices,
        'solutii_choices': SolutieDosar._meta.get_field('tip_solutie').choices,
        'page_obj': page_obj,
        'total_rezultate': paginator.count, # Numărul total real de rezultate găsite
        'query_string': query_string, # Pentru a păstra filtrele și coloanele selectate în link-urile de paginare
    }
    
    return render(request, 'cases/rapoarte.html', context)

@login_required
def harta_infractionalitatii(request):
    # Luăm doar infracțiunile care au coordonatele completate
    infractiuni = Infractiune.objects.filter(latitudine__isnull=False, longitudine__isnull=False)
    
    # Le formatăm într-o listă de dicționare pentru a le putea citi în JavaScript
    date_harta = []
    # luăm datele cu numele lor reale din baza de date:
    for inf in infractiuni:
        date_harta.append({
            'lat': inf.latitudine,
            'lng': inf.longitudine,
            'dosar': inf.dosar.numar_unic,
            'dosar_id': inf.dosar.pk,
            'act_normativ': inf.get_act_normativ_display() or "Necunoscut",
            'incadrare': inf.incadrare_juridica if inf.incadrare_juridica else "Fără încadrare", # <--- NOU
            'articol': inf.articol if inf.articol else "-", # <--- NOU
            'adresa': inf.adresa_comiterii,
            'data': inf.data_comiterii.isoformat() if inf.data_comiterii else None
        })
        
    context = {
        # Transformăm lista în format JSON pentru HTML
        'date_harta_json': json.dumps(date_harta),
        'tile_server_url': settings.MAP_TILE_SERVER_URL, # Trimitem adresa hărții
    }
    
    return render(request, 'cases/harta.html', context)

# Calea către executabilul Tesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

@csrf_exempt  # Dezactivăm temporar protecția CSRF doar pentru acest test
def test_ocr_api(request):
    if request.method == 'POST' and request.FILES.get('imagine_buletin'):
        fisier_imagine = request.FILES['imagine_buletin']
        
        try:
            img = Image.open(fisier_imagine)
            width, height = img.size
            img = img.resize((width * 2, height * 2), Image.Resampling.LANCZOS)
            img = img.convert('L')
            img = ImageEnhance.Contrast(img).enhance(2.5)
            img = ImageEnhance.Sharpness(img).enhance(2.0)
            
            config_tesseract = r'--oem 3 --psm 6'
            text_brut = pytesseract.image_to_string(img, lang='ron', config=config_tesseract)
            
            # Datele pe care vrem să le completăm
            date_extrase = {
                'cnp': '',
                'nume_complet': '',
                'adresa': '',
                'serie': '',
                'numar': ''
            }
            
            # 1. Căutăm CNP-ul
            cnp_match = re.search(r'\b[1-9]\d{12}\b', text_brut)
            if cnp_match:
                date_extrase['cnp'] = cnp_match.group(0)

            # 2 şi 3. Căutăm Seria și Numărul (Împreună)
            # Caută (opțional) eticheta "ID:", urmată de 2 litere (Grupul 1), 
            # spații opționale, și fix 6 cifre (Grupul 2).
            serie_numar_match = re.search(r'(?:ID:\s*)?([A-Z]{2})\s*(\d{6})', text_brut, re.IGNORECASE)
            
            if serie_numar_match:
                date_extrase['serie'] = serie_numar_match.group(1).upper()  
                date_extrase['numar'] = serie_numar_match.group(2)          
                
            # 2. Funcție de curățare a numelor (CU REGULA VOCALEI)
            def curata_nume(text):
                if not text: return ""
                # Extragem cuvintele de minim 2 litere (majuscule)
                cuvinte_brute = re.findall(r'\b[A-ZĂÎÂȘȚ\-]{2,}\b', text.upper())
                cuvinte_valide = []
                
                for cuvant in cuvinte_brute:
                    # Dacă acel cuvânt are cel puțin o vocală, e nume valid
                    # Astfel, construcţii ca 'LL', 'XX', 'M' nu vor fi acceptate
                    if re.search(r'[AEIOUĂÎÂ]', cuvant):
                        cuvinte_valide.append(cuvant)
                        
                return " ".join(cuvinte_valide)

            nume_match = re.search(r'NUME:\s*([^\n]+)', text_brut, re.IGNORECASE)
            prenume_match = re.search(r'PRENUME:\s*([^\n]+)', text_brut, re.IGNORECASE)
            
            nume_final = ""
            if nume_match:
                nume_final += curata_nume(nume_match.group(1))
            if prenume_match:
                nume_final += " " + curata_nume(prenume_match.group(1))
                
            date_extrase['nume_complet'] = nume_final.strip()
            
            # 3. Căutăm și curățăm Adresa/Domiciliul
            adresa_match = re.search(r'(?:ADRESA|DOMICILIU):\s*([^\n]+)', text_brut, re.IGNORECASE)
            if adresa_match:
                adresa_bruta = adresa_match.group(1).strip()
                # Eliminăm literele mici izolate 
                adresa_curata = re.sub(r'\b[a-z]\b', '', adresa_bruta)
                # Eliminăm spațiile duble rezultate
                adresa_curata = re.sub(r'\s+', ' ', adresa_curata).strip()
                
                date_extrase['adresa'] = adresa_curata

            return JsonResponse({
                'status': 'success', 
                'date_structurate': date_extrase
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'mesaj': str(e)})
            
    return JsonResponse({'status': 'error', 'mesaj': 'Metodă invalidă.'})

# Funcția care afișează pagina HTML pe ecran
@login_required
def pagina_test_ocr(request):
    return render(request, 'cases/test_ocr.html')

# === GRAF RELAŢIONAL ===
# API apelat de JavaScript pentru a obține datele în formatul necesar pentru vizualizarea grafică
# extrage datele, rezolvă problema duplicatelor și construiește structura matematică (Noduri și Muchii)
@login_required
def date_graf_relational(request):
    """
    Acest API returnează datele întregii baze de date 
    formatate pentru biblioteca de grafuri Vis.js
    """
    nodes = []
    edges = []
    
    # ==========================================
    # 1. GENERĂM NODURILE PENTRU DOSARE
    # ==========================================
    dosare = Dosar.objects.all().prefetch_related('stadii_cercetare')
    for d in dosare:
        stadiu = d.stadiu_curent
        nume_stadiu = stadiu.get_tip_stadiu_display() if stadiu else "În lucru"

        # Verificăm dacă are o soluție
        solutie = d.solutie_curenta
        text_solutie = f"\nSoluție: {solutie.get_tip_solutie_display()}" if solutie else ""
        
        nodes.append({
            'id': f'dosar_{d.id}',
            'label': f'Dosar\n{d.numar_unic}',
            'group': 'dosar',
            'title': f'Dosar: {d.numar_unic}\nStadiu: {nume_stadiu}{text_solutie}',
            'url_dosar': reverse('cases:detalii_dosar', args=[d.pk]) # <--- Trimitem link-ul dosarului
        })
        
    # ==========================================
    # 2. GENERĂM NODURILE PENTRU PERSOANE ȘI LINIILE (MUCHIILE)
    # ==========================================
    parti = ParteImplicata.objects.select_related('dosar').all()
    
    # Folosim un dicționar pentru a păstra persoanele și toate calitățile lor
    # Cheia va fi ID-ul unic (CNP sau Nume), caloarea va fi un dicționar cu Nume și Lista de calități
    persoane_unice = {}
    
    for parte in parti:
        # Identificarea unică a unei persoane. Dacă are CNP, e unic.
        # Dacă nu are CNP, folosim numele lui transformat cu majuscule și fără spații (ex: ION_POPESCU)
        if parte.cnp:
            nod_id = f'pers_cnp_{parte.cnp}'
            label_nod = f'{parte.nume_complet}\n({parte.cnp})'
        else:
            nume_curat = parte.nume_complet.strip().upper().replace(" ", "_")
            nod_id = f'pers_nume_{nume_curat}'
            label_nod = parte.nume_complet
            
        calitate = parte.get_calitate_procesuala_display()
            
        # Dacă persoana nu există încă pe hartă, o adăugăm în dicționar
        if nod_id not in persoane_unice:
            persoane_unice[nod_id] = {
                'id': nod_id,
                'label': label_nod,
                'nume_complet': parte.nume_complet,
                'calitati': {calitate} # Folosim un Set pentru calități unice (ex: să nu apară Suspect de 2 ori)
            }
        else:
            # Dacă există deja, doar îi adăugăm noua calitate
            persoane_unice[nod_id]['calitati'].add(calitate)
            
        # Adăugăm linia între persoană și dosar
        edges.append({
            'from': nod_id,
            'to': f'dosar_{parte.dosar.id}',
            'label': calitate,
            'font': {'align': 'middle'},
            'color': {'color': '#a3a3a3', 'highlight': '#ff4b4b'}
        })

    # Construim lista finală de noduri pentru persoane
    for pers_id, date_pers in persoane_unice.items():
        # Creăm un text cu toate calitățile pentru Tooltip (la hover)
        lista_calitati_html = ", ".join(date_pers['calitati'])
        
        nodes.append({
            'id': date_pers['id'],
            'label': date_pers['label'],
            'group': 'persoana',
            'title': f"{date_pers['nume_complet']}\nCalitate: {lista_calitati_html}",
            # Salvăm un câmp invizibil cu calitățile ca să le citească JavaScript pentru panoul lateral
            'calitati_js': lista_calitati_html 
        })
        
    return JsonResponse({'nodes': nodes, 'edges': edges})

# Funcția simplă care va deschide pagina HTML a Grafului
@login_required
def graf_relational(request):
    return render(request, 'cases/graf_relational.html')

@login_required
def genereaza_act(request, pk):
    dosar = get_object_or_404(Dosar, pk=pk)

    if not dosar.are_drepturi_editare(request.user):
        raise PermissionDenied("Nu ai permisiunea de a edita acest dosar.")
    
    if request.method == 'POST':
        tip_act = request.POST.get('tip_act')
        
        # --- 1. SETĂM EMITENTUL ȘI VARIABILELE INDIVIDUALE ---
        alegere_emitent = request.POST.get('emitent', 'politie') 
        
        # Reguli stricte de competență
        if tip_act in ['ord_clasare', 'ord_renuntare']:
            alegere_emitent = 'procuror'
        elif tip_act in ['ref_clasare', 'ref_renuntare']:
            alegere_emitent = 'politie'
            
        nume_user = request.user.get_full_name() or request.user.username
        unitate = getattr(request.user, 'unitate', 'Unitate nespecificată')
        grad = getattr(request.user, 'grad_profesional', '')
        
        if alegere_emitent == 'procuror':
            functie_antet = "Procuror"
            nume_semnatar = nume_user
        else:
            functie_antet = "Organ de cercetare penală al poliției judiciare"
            nume_semnatar = f"{grad} {nume_user}"

        # --- 2. LOGICA PENTRU FIECARE ACT ȘI ALEGEREA ȘABLONULUI ---
        context_doc = {}
        template_name = ""
        nume_fisier = ""
        
        if tip_act == 'citatie':
            id_persoana = request.POST.get('persoana_citata')
            parte = get_object_or_404(ParteImplicata, pk=id_persoana)
            
            # FORMATĂM DATA ȘI ORA PENTRU CITAȚIE
            data_ora_raw = request.POST.get('data_ora')
            data_ora_formatata = data_ora_raw 
            if data_ora_raw:
                try:
                    dt_obj = datetime.strptime(data_ora_raw, '%Y-%m-%dT%H:%M')
                    data_ora_formatata = dt_obj.strftime('%d.%m.%Y, ora %H:%M')

                    # CREAREA AUTOMATĂ A TERMENULUI DE AUDIERE ÎN SECŢIUNEA DE TERMENE
                    TermenProcedural.objects.create(
                        dosar=dosar,
                        tip_termen='AUDIERE',
                        data_limita=dt_obj.date(),
                        detalii=f"Citat: {parte.nume_complet} pentru {request.POST.get('scop')}"
                    )
                except ValueError:
                    pass
            
            if alegere_emitent == 'procuror':
                template_name = 'citatie_procuror_template.docx'
            else:
                template_name = 'citatie_politie_template.docx'
                
            nume_fisier = f"Citatie_{parte.nume_complet.replace(' ', '_')}_{dosar.numar_unic}.docx"
            
            context_doc = {
                'numar_dosar': dosar.numar_unic,
                'nume_persoana': parte.nume_complet,
                'calitate': parte.get_calitate_procesuala_display() if hasattr(parte, 'get_calitate_procesuala_display') else parte.calitate_procesuala,
                'adresa': parte.adresa if parte.adresa else "Adresă necunoscută",
                'data_ora': data_ora_formatata, 
                'scop': request.POST.get('scop'),
                'functie': functie_antet,
                'nume_semnatar': nume_semnatar,
                'unitate': unitate
            }
            
        elif tip_act in ['oiup', 'ord_clasare', 'ref_clasare', 'ord_renuntare', 'ref_renuntare']:
            # Formatăm data emiterii actului
            data_raw = request.POST.get('data_actului')
            data_formatata = "____/____/________"
            if data_raw:
                dt = datetime.strptime(data_raw, '%Y-%m-%d')
                data_formatata = dt.strftime('%d.%m.%Y')
            
            # --- DATE SPECIFICE DOSARULUI ---
            data_inreg = dosar.data_inregistrarii.strftime('%d.%m.%Y') if dosar.data_inregistrarii else "---"
            
            infr_raw = getattr(dosar, 'infractiune_cercetata', 'faptele reclamate')
            if infr_raw and len(infr_raw) > 0:
                infr_cercetata = infr_raw[0].lower() + infr_raw[1:]
            else:
                infr_cercetata = "faptele reclamate"
            
            prima_infractiune = dosar.infractiuni.first()
            if prima_infractiune:
                incad_raw = prima_infractiune.incadrare_juridica
                if incad_raw and len(incad_raw) > 0:
                    incadrare = incad_raw[0].lower() + incad_raw[1:]
                else:
                    incadrare = "nespecificată"
                    
                articol = prima_infractiune.articol
                act_normativ = prima_infractiune.get_act_normativ_display() if hasattr(prima_infractiune, 'get_act_normativ_display') else prima_infractiune.act_normativ
            else:
                incadrare = "nespecificată"
                articol = "---"
                act_normativ = "---"

            # Lista persoanelor alese pentru comunicare
            ids_comunicare = request.POST.getlist('persoane_comunicare') 
            parti_selectate = ParteImplicata.objects.filter(id__in=ids_comunicare)
            lista_parti = [{'nume_complet': p.nume_complet} for p in parti_selectate]

            # --- CONTEXTUL COMUN ---
            context_doc = {
                'numar_dosar': dosar.numar_unic,
                'data_actului': data_formatata,
                'temei': request.POST.get('temei_clasare', ''), 
                'functie': functie_antet,
                'nume_semnatar': nume_semnatar,
                'unitate': unitate,
                'data_inregistrarii': data_inreg,
                'infractiune_cercetata': infr_cercetata,
                'incadrare_juridica': incadrare,
                'articol': articol,
                'act_normativ': act_normativ,
                'lista_comunicare': lista_parti 
            }
            
            # ---> BLOCUL CARE LIPSEA: SETAREA ȘABLOANELOR <---
            if tip_act == 'oiup':
                if alegere_emitent == 'procuror':
                    template_name = 'oiup_procuror_template.docx'
                else:
                    template_name = 'oiup_politie_template.docx'
                nume_fisier = f"OIUP_{dosar.numar_unic}.docx"
                
            elif tip_act == 'ord_clasare':
                template_name = 'ordonanta_clasare.docx'
                nume_fisier = f"Ordonanta_Clasare_{dosar.numar_unic}.docx"
                
            elif tip_act == 'ref_clasare':
                template_name = 'clasare_ref_template.docx'
                nume_fisier = f"Referat_Clasare_{dosar.numar_unic}.docx"
                
            elif tip_act == 'ord_renuntare':
                template_name = 'renuntare_ord_template.docx'
                nume_fisier = f"Ordonanta_Renuntare_{dosar.numar_unic}.docx"
                
            elif tip_act == 'ref_renuntare':
                template_name = 'renuntare_ref_template.docx'
                nume_fisier = f"Referat_Renuntare_{dosar.numar_unic}.docx"

        # --- 3. GENERAREA FIZICĂ A DOCUMENTULUI ȘI SALVAREA ---
        if template_name:
            template_path = os.path.join(settings.BASE_DIR, 'cases', 'templates', 'acte', template_name)
            
            if not os.path.exists(template_path):
                messages.error(request, f'Eroare: Șablonul {template_name} nu a fost găsit pe server.')
                return redirect('cases:detalii_dosar', pk=dosar.pk)

            doc = DocxTemplate(template_path)
            doc.render(context_doc)
            
            file_stream = io.BytesIO()
            doc.save(file_stream)
            
            # --- NOU: SALVAREA ÎN BAZA DE DATE ---
            # Facem un dicționar pentru a transforma codul actului într-un text lizibil
            nume_tipuri_acte = {
                'citatie': 'Citație',
                'oiup': 'Ordonanță Începere UP',
                'ord_clasare': 'Ordonanță Clasare',
                'ref_clasare': 'Referat Clasare',
                'ord_renuntare': 'Ordonanță Renunțare',
                'ref_renuntare': 'Referat Renunțare',
            }
            tip_frumos = nume_tipuri_acte.get(tip_act, tip_act.upper())
            
            # Curățăm titlul ca să nu arate urât în tabelă (ex: "Citatie Popescu_Ion_123_P_2024" -> "Citatie Popescu Ion 123 P 2024")
            titlu_curat = nume_fisier.replace('.docx', '').replace('_', ' ')

            # Creăm și salvăm obiectul
            # (Presupun că ai importat modelul ActUrmarire sus de tot)
            act = ActUrmarire(
                dosar=dosar,
                tip=tip_frumos[:50], 
                titlu=titlu_curat,
                autor=request.user,
                data_documentului=datetime.today().date(),
            )
            
            # act.fisier.save() face automat și scrierea fișierului pe disc (în media/) și act.save() în baza de date!
            act.fisier.save(nume_fisier.replace("/", "_"), ContentFile(file_stream.getvalue()))
            
            # Mesaj de succes care va apărea când utilizatorul dă refresh la pagina dosarului (fiindcă fișierul se descarcă automat, browserul nu dă refresh instant)
            messages.success(request, f'{tip_frumos} a fost generat(ă) și salvat(ă) în arhiva dosarului.')

            # Resetăm pointerul stream-ului pentru a-l trimite și spre descărcare
            file_stream.seek(0)
            response = HttpResponse(file_stream.read(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = f'attachment; filename="{nume_fisier.replace("/", "_")}"'
            
            return response

    messages.error(request, 'A apărut o eroare la generarea actului. Verificați datele.')
    return redirect('cases:detalii_dosar', pk=dosar.pk)

@login_required
def semneaza_act(request, pk_act):
    # 1. Găsim actul (presupunând că modelul se numește ActUrmarire)
    act = get_object_or_404(ActUrmarire, pk=pk_act)
    dosar = act.dosar
    user = request.user
    
    # 2. Verificări de siguranță
    if not dosar.are_drepturi_editare(user):
        raise PermissionDenied("Nu ai permisiunea de a semna în acest dosar.")
        
    if not act.fisier.name.lower().endswith('.pdf'):
        messages.error(request, "Eroare: Doar documentele în format PDF pot fi semnate digital.")
        return redirect('cases:detalii_dosar', pk=dosar.pk)
        
    if not user.cheie_privata_criptata or not user.certificat_pem:
        messages.error(request, "Nu ai o identitate digitală generată. Contactează administratorul.")
        return redirect('cases:detalii_dosar', pk=dosar.pk)

    if request.method == 'POST':
        try:
            # 3. DECRIPTĂM CHEIA PRIVATĂ (Folosind cheia Fernet din settings)
            f = Fernet(settings.ENCRYPTION_KEY)
            private_key_pem = f.decrypt(user.cheie_privata_criptata)
            
            # --- MODIFICAREA ESTE AICI ---
# 4. PREGĂTIM OBIECTELE CRIPTOGRAFICE (EXCLUSIV CU asn1crypto)
            
            # a) Încărcăm cheia privată
            # private_key_pem a fost deja decriptat cu Fernet mai sus
            _, _, priv_der_bytes = pem.unarmor(private_key_pem)
            private_key = asn1_keys.PrivateKeyInfo.load(priv_der_bytes)
            
            # b) Încărcăm certificatul public
            cert_bytes = user.certificat_pem.encode('utf-8')
            _, _, cert_der_bytes = pem.unarmor(cert_bytes)
            cert = asn1_x509.Certificate.load(cert_der_bytes)
            
            # c) Inițializăm semnatarul pyHanko
            signer = signers.SimpleSigner(
                signing_cert=cert,
                signing_key=private_key,
                cert_registry=None
            )
            # --- SFÂRȘIT MODIFICARE ---
            
            # 5. DESCHIDEM PDF-UL ȘI PREGĂTIM SCRIEREA
# 5. DESCHIDEM PDF-UL ÎN MEMORIE (Citire + Scriere)
            pdf_stream = act.fisier.open('rb')
            pdf_bytes = pdf_stream.read()
            pdf_stream.close() # Închidem stream-ul original
            
            # Punem biții într-un buffer nou, unde pyHanko va putea și citi, dar și scrie semnătura!
            in_out_buffer = io.BytesIO(pdf_bytes)
            pdf_writer = IncrementalPdfFileWriter(in_out_buffer, strict=False)
            
            # 6. APLICĂM SEMNĂTURA CRIPTOGRAFICĂ
            # 6. APLICĂM SEMNĂTURA CRIPTOGRAFICĂ (CU ȘTAMPILĂ VIZUALĂ)
            nume_camp = f'Semnatura_{user.username}'
            
            # a) Creăm un "dreptunghi" vizibil pe prima pagină
            # Sistemul de coordonate: (X_stânga, Y_jos, X_dreapta, Y_sus)
            # Punctul 0,0 este în colțul stânga-jos al paginii.
            # O pagină A4 are cam 595 x 842 puncte. (300, 50, 550, 130) îl va pune în dreapta-jos.
            append_signature_field(
                pdf_writer,
                SigFieldSpec(
                    sig_field_name=nume_camp,
                    on_page=-1, # 0 = prima pagină, -1 = ultima pagină
                    box=(320, 40, 550, 120) 
                )
            )
            
            # b) Definim cum arată textul ștampilei
            # Folosim Python (f-strings) pentru a injecta Numele și Motivul. 
            # Lăsăm doar %(ts)s pentru ca pyHanko să pună secunda exactă a semnăturii.
            nume_afisare = user.get_full_name() or user.username
            motiv_semnare = "Semnatura digitala aprobata"
            
            stil_stampila = text.TextStampStyle(
                stamp_text=f'SEMNAT DIGITAL\nSemnatar: {nume_afisare}\nData: %(ts)s\nMotiv: {motiv_semnare}'
            )
            
            # c) Inițializăm procesul de semnare (atașând și stilul vizual creat)
            pdf_signer = signers.PdfSigner(
                signers.PdfSignatureMetadata(
                    field_name=nume_camp, 
                    reason="Semnătură digitală aprobată",
                    location=getattr(user, 'unitate', "Sistem Management Dosare")
                ),
                signer=signer,
                stamp_style=stil_stampila
            )
            
# d) Scriem semnătura pe fișier
            pdf_signer.sign_pdf(pdf_writer, in_place=True)
            
            # 7. SALVĂM VERSIUNEA SEMNATĂ (FĂRĂ SĂ DISTRUGEM ORIGINALUL)
            in_out_buffer.seek(0)
            
            # Extragem numele original (ex: "ordonanta_123") și extensia (".pdf")
            nume_original, extensie = os.path.splitext(os.path.basename(act.fisier.name))
            nume_fisier_semnat = f"{nume_original}_semnat{extensie}"
            
            # Salvăm documentul în noul câmp 'fisier_semnat'
            act.fisier_semnat.save(nume_fisier_semnat, ContentFile(in_out_buffer.read()), save=False)
            
            # Actualizăm noile câmpuri din model
            act.este_semnat = True
            act.semnat_de = user
            act.data_semnarii = timezone.now()
            # Am eliminat complet acel hack cu act.titlu = "[SEMNAT]" !
            
            act.save()
            
            messages.success(request, f'Documentul "{act.titlu}" a fost semnat digital cu succes!')
            
        except Exception as e:
            # Jurnalizăm eroarea în fundal pentru administratori
            logger.error(
                f"Eroare semnare act {pk_act} de user {user.username}: "
                f"{type(e).__name__}: {str(e)}", 
                exc_info=True
            )
            # Afișăm un mesaj elegant utilizatorului
            messages.error(
                request, 
                "A apărut o eroare la semnarea documentului. Verificați că fișierul PDF este valid și reîncercați."
            )
            
    return redirect('cases:detalii_dosar', pk=dosar.pk)